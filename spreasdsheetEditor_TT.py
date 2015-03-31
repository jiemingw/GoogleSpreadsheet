import sys
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter

def main():
    arg=sys.argv[1]
    filename=str(arg)
    print('filename is', filename)
    wb = load_workbook(filename = filename,data_only=True)
    new_sheet= wb.create_sheet()
    new_sheet.title="Labor_Rates_Table"
    entryNumber=1;

    sheet = wb['PM_CM']
    for col_idx in range(5, 22):
        for row in range(13, 41):
            entryNumber=handle_cell(sheet,new_sheet,entryNumber,col_idx,row,7,'B',5,6,8,9)

    sheet0 = wb['A_E']
    #print(sheet1, filename)
    for col_idx in range(5, 64):
        for row in range(13, 66):
            entryNumber=handle_cell(sheet0,new_sheet,entryNumber,col_idx,row,7,'B',5,6,8,9)

    sheet1 = wb['Commissioning']
    for col_idx in range(5, 22):
        for row in range(13, 26):
            entryNumber=handle_cell(sheet1,new_sheet,entryNumber,col_idx,row,7,'B',5,6,8,9)

    sheet2 = wb['GC']
    for col_idx in range(5, 29):
        for row in range(13, 46):
            entryNumber=handle_cell(sheet2,new_sheet,entryNumber,col_idx,row,7,'B',5,6,8,9)

    sheet3 = wb['Carbon']
    for col_idx in range(4, 8):
        for row in range(13, 31):
            entryNumber=handle_cell(sheet3,new_sheet,entryNumber,col_idx,row,7,'B',5,6,8,9)

    sheet4 = wb['Security']
    for col_idx in range(5, 9):
        for row in range(13, 60):
            entryNumber=handle_cell(sheet4,new_sheet,entryNumber,col_idx,row,7,'B',5,6,8,9)

    wb.save(filename = 'LaborDone.xlsx')

def handle_cell(sheet_ranges,raw_data_sheet,entryNumber,col_idx,row,discipline_row, role_column, date_row, company_row, region_row, conversion_row):

    col = get_column_letter(col_idx)
    rate=(sheet_ranges.cell('%s%s'%(col, row)).value)
    discipline=(sheet_ranges.cell('%s%s'%(col, discipline_row)).value)
    role=(sheet_ranges.cell('%s%s'%(role_column, row)).value)
    company=(sheet_ranges.cell('%s%s'%(col, company_row)).value)
    region=(sheet_ranges.cell('%s%s'%(col, region_row)).value)
    date=(sheet_ranges.cell('%s%s'%(col, date_row)).value)
    conversion=(sheet_ranges.cell('%s%s'%(col, conversion_row)).value)

    if rate is not None:
        entryNumber=entryNumber+1
        #if type(rate) is int:
        raw_data_sheet.cell('%s%s'%('D', entryNumber)).value=rate;
        #if type(rate) is str:
        #    afterEqualSign=rate.split('=')[1]
        #    stringRate=afterEqualSign.split('/')[0]
        #    wei_sheet.cell('%s%s'%('D', entryNumber)).value=float(stringRate)/conversion;
        raw_data_sheet.cell('%s%s'%('B', entryNumber)).value=discipline;
        raw_data_sheet.cell('%s%s'%('C', entryNumber)).value=role;
        raw_data_sheet.cell('%s%s'%('A', entryNumber)).value=company;
        raw_data_sheet.cell('%s%s'%('E', entryNumber)).value=region;
        raw_data_sheet.cell('%s%s'%('G', entryNumber)).value=date;
        if conversion is not None:
            raw_data_sheet.cell('%s%s'%('F', entryNumber)).value=conversion;
        else:
            raw_data_sheet.cell('%s%s'%('F', entryNumber)).value=1;

    return entryNumber

if __name__ == "__main__":
    main()
