import sys
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter

def main():
    arg=sys.argv[1]
    filename=str(arg)
    print('filename is', filename)
    wb = load_workbook(filename = filename,data_only=True)
    sheet_ranges = wb['A_E']
    wei_sheet= wb.create_sheet()
    entryNumber=1;
    for col_idx in range(5, 64):
        col = get_column_letter(col_idx)
        for row in range(13, 66):
            rate=(sheet_ranges.cell('%s%s'%(col, row)).value)
            discipline=(sheet_ranges.cell('%s%s'%(col, 7)).value)
            role=(sheet_ranges.cell('%s%s'%('B', row)).value)
            company=(sheet_ranges.cell('%s%s'%(col, 6)).value)
            conversion=(sheet_ranges.cell('%s%s'%(col, 9)).value)

            if rate is not None:
                entryNumber=entryNumber+1
                #if type(rate) is int:
                wei_sheet.cell('%s%s'%('A', entryNumber)).value=rate;
                #if type(rate) is str:
                #    afterEqualSign=rate.split('=')[1]
                #    stringRate=afterEqualSign.split('/')[0]
                #    wei_sheet.cell('%s%s'%('A', entryNumber)).value=float(stringRate)/conversion;
                wei_sheet.cell('%s%s'%('B', entryNumber)).value=discipline;
                wei_sheet.cell('%s%s'%('C', entryNumber)).value=role;
                wei_sheet.cell('%s%s'%('D', entryNumber)).value=company;

    wb.save(filename = 'Wei.xlsx')


if __name__ == "__main__":
    main()
